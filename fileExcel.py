import openpyxl

import os

os.chdir('C:\\Users\\Asus\\Desktop\\NER')

workbook = openpyxl.load_workbook('ner.xlsx')
print(workbook)

sheet = workbook.get_sheet_by_name('Sheet1')

print(workbook.get_sheet_names())
cell = sheet['A1']
print(cell.value)

# sheet.cell(row =1, column = 2)
# cell = sheet['B1']
"""
for i in range(1, 9):
    if str(sheet.cell(row=i, column=1).value) != 'None':
        print(i, sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value)

"""
sheet_obj = workbook.active
m_row = sheet_obj.max_row

# Loop will print all values
# of first column
# for i in range(1, m_row + 1):
#  print(i, sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value)

max_col = sheet_obj.max_column
# Will print a particular row value and column
print('\nWill print a particular row value and column')
for i in range(1, m_row + 1):
    for j in range(1, max_col + 1):
        cell_obj = sheet_obj.cell(row=i, column=j)
        print(cell_obj.value, end= ' ')
    print('\n')

# creating new sheet in excel
sheet2 = workbook.create_sheet()
sheet2.title = 'My new sheet Name'
print(workbook.get_sheet_names())
workbook.save('ner1.xlsx')